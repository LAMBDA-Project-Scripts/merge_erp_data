import os
import re
import statistics
from collections import defaultdict
from dataclasses import dataclass
from enum import Enum
from gpt2_tools import LLM_Tool
from openpyxl import load_workbook


# Class representing all the information we want to collect
@dataclass
class Record:
    text : str
    full_id : str
    lex : str
    condition : str
    animacy : str
    surprisal : float
    perplexity : float
    group : str
    ambiguity : str
    ref_judgement : str

    def is_consistent(self, allow_missing=None):
        """Checks consistency for a specific record.

        Parameters
        ----------
        allow_missing : list(str)
            List of fields that we allow to be inconsistent.
        """
        if "text" not in allow_missing:
            assert self.text is not None and len(self.text) > 0, \
                "Text is missing"
        if "full_id" not in allow_missing:
            assert self.full_id is not None and len(self.full_id) > 0, \
                "Full_id is missing"
        if "lex" not in allow_missing:
            assert self.lex is not None and len(self.lex) > 0, \
                "Lex is missing"
        if "condition" not in allow_missing:
            assert self.condition is not None and len(self.condition) > 0, \
                "Condition is missing"
        if "animacy" not in allow_missing:
            assert self.animacy is not None and self.animacy < 2, \
                "Animacy is missing"
        if "surprisal" not in allow_missing:
            assert self.surprisal is not None and self.surprisal > 0, \
                "Surprisal is missing"
        if "perplexity" not in allow_missing:
            assert self.perplexity is not None and self.perplexity > 0, \
                "Perplexity is missing"
        if "group" not in allow_missing:
            assert self.group is not None and len(self.group) > 0, \
                "Group is missing"
        if "ambiguity" not in self.ambiguity:
            assert self.ambiguity is not None and len(self.ambiguity) > 0, \
                "Ambiguity is missing"
        if "ref_judgement" not in allow_missing:
            assert self.ref_judgement is not None and \
                isinstance(self.ref_judgement, Antecedent), \
                "Referential judgement is missing"
		

class Antecedent(Enum):
	FIRST = 1
	SECOND = 2
	BOTH = 3


def calculate_antecedent(first_val, second_val):
	assert 1<=first_val<=5, f"Invalid value {first_val} for first_val"
	assert 1<=second_val<=5, f"Invalid value {second_val} for second_val"
	values = {(1,3): Antecedent.FIRST,
			  (1,4): Antecedent.FIRST,
			  (1,5): Antecedent.FIRST,
			  (2,4): Antecedent.FIRST,
			  (2,5): Antecedent.FIRST,
			  (3,1): Antecedent.SECOND,
			  (3,2): Antecedent.SECOND, # This doesn't hold up there
			  (3,5): Antecedent.FIRST,
			  (4,1): Antecedent.SECOND,
			  (4,2): Antecedent.SECOND,
			  (5,1): Antecedent.SECOND,
			  (5,2): Antecedent.SECOND,
			  (5,3): Antecedent.SECOND}
	try:
		return values[(first_val, second_val)]
	except KeyError:
		return Antecedent.BOTH


if __name__ == '__main__':
	# Data that will be put together from multiple sources
	worksheet_data = []
	# All sources of data
	source_excel = os.path.join('data', 'Items_final_all_22_11_2022.xlsx')
	animacy_filename = os.path.join('data',
									'proref_condition_coding_itemlist_fixed.txt')
	structure_pattern_filename = os.path.join('data',
									          'structure_pattern_lex_condition.txt')
	proref_data_filename = os.path.join('data', 
	                                    'proref_erp_data_structure_example.txt')
	proref_questionnaire_filename = os.path.join('data', 'Proref_Questionnaire_Data_2023-10-27.xlsx')
	outfile = 'results.csv'

	# GPT-2 measures
	llmtool = LLM_Tool()

	# First, read data from the Items worksheet.
	workbook = load_workbook(filename=source_excel)
	worksheet = workbook['Items_all_corrct']

	# We hard-code the number of rows for now, but hopefully we can
	# detect that automatically in the future.
	for row in range(1, 481):
		item_read = worksheet.cell(row=row, column=3).value + ' ' + worksheet.cell(row=row, column=4).value
		code = str(worksheet.cell(row=row, column=2).value)
		if item_read is not None:
			# First, process the text in column D.
			# We start by removing everything after the last "und ..."
			elements = item_read.split(' und')
			surprisal_text = ' '.join(elements[:-1])
			# Next, remove special characters used to mark parts of the
			# sentence
			for character in ['§', '!', '*', '`', '&', '_']:
				surprisal_text = surprisal_text.replace(character, ' ')
			# Finally, collapse all multiple spaces into a single one
			surprisal_text = ' '.join(surprisal_text.split())

			# Now, the codes in column B
			lex = code[0:-5]
			second_code = code[-5]
			cond_pos = code[-4:-1]
			# We are currently not using this one
			behavioral_response = code[-1:]
			
			# Finally, surprisal for the input text
			tokens = llmtool.get_tokenizer().encode(surprisal_text, return_tensors='pt')
			surprisal = llmtool.get_text_surprisal(tokens)
			perplexity = llmtool.get_text_perplexity(tokens)

			worksheet_data.append(Record(text=surprisal_text,
			                             full_id=code,
			                             lex=lex,
			                             condition=cond_pos,
			                             animacy=None,
			                             surprisal=surprisal,
			                             perplexity=perplexity,
			                             group=None,
			                             ambiguity=None,
			                             ref_judgement=None
			                             ))
	assert len(worksheet_data) == 480, "Not enough data was read properly"

	# We now read the animacy data and put it together with the data
	# we collected from the spreadsheet
	with open(animacy_filename, 'r') as fp:
		# Skip the header
		next(fp)
		for idx, line in enumerate(fp):
			# Read the animacy data from the animacy text file,
			# performing first some data quality checks
			fields = re.split(' |\t', line.strip())
			assert fields[0] == worksheet_data[idx].lex, \
			       "Data field 'lex' is not properly paired"
			assert fields[1] == worksheet_data[idx].condition, \
			       "Data field 'condition' is not properly paired"
			# Update the value of the 'animacy' field
			assert fields[2] in {'i', 'a'}, \
					"Data field 'animacy' has an invalid value"
			worksheet_data[idx].animacy = 1 if fields[2] == 'a' else 0

	# Next is the data regarding ambiguity and subgroups
	with open(structure_pattern_filename, 'r') as fp:
		# Skip the header
		next(fp)
		for idx, line in enumerate(fp):
			fields = re.split(' |\t', line.strip())
			assert fields[1] == worksheet_data[idx].lex, \
			       "Data field 'lex' is not properly paired"
			cond_data = fields[0].split('_')
			group = '_'.join(cond_data[0:-2])
			ambig = cond_data[-2]
			worksheet_data[idx].group = group
			# We convert 'amb/unamb' into 'ambig/unambig'
			worksheet_data[idx].ambiguity = ambig + 'ig'

	# Next step: collect referential judgements for *some* fields
	proref_item_to_rating = defaultdict(list)
	workbook = load_workbook(filename=proref_questionnaire_filename)
	for list_num in range(1, 5):
		worksheet = workbook[f'list_{list_num}']
		row = 2
		done = False
		while not done:
			try:
				item_num = worksheet.cell(row=row, column=6).value
				rating1 = int(worksheet.cell(row=row, column=10).value)
				rating2 = int(worksheet.cell(row=row, column=11).value)
				proref_item_to_rating[(item_num, 1)].append(rating1)
				proref_item_to_rating[(item_num, 2)].append(rating2)
			except TypeError:
				# Done iterating this workbook
				done = True
			except ValueError:
				# Some records just say "N/A" instead of a rating
				pass
			row += 1
	for record in worksheet_data:
		try:
			mean_rating1 = statistics.mean(proref_item_to_rating[(int(record.full_id[:-1]), 1)])
			mean_rating2 = statistics.mean(proref_item_to_rating[(int(record.full_id[:-1]), 2)])
			record.ref_judgement = calculate_antecedent(round(mean_rating1), round(mean_rating2))
		except statistics.StatisticsError:
			pass

	# Final check: all data has been filled for all fields
	keys_to_record = dict()
	for record in worksheet_data:
		# Check that the record is consistent, except for ref_judgement
		# which we know can be missing
		record.is_consistent(allow_missing=["ref_judgement"])
		# The key we use is (subject, group, ambiguity)
		key = ('S' + record.lex, record.group, record.ambiguity)
		if record.ref_judgement is not None and isinstance(record.ref_judgement, Antecedent):
			# We have a referential judgement value
			keys_to_record[key] = (record.animacy, record.surprisal, record.perplexity, record.ref_judgement.name)
		else:
			# We don't have a referential judgement value
			keys_to_record[key] = (record.animacy, record.surprisal, record.perplexity, 'N/A')
		
	# Unify all data
	first = True
	with open(proref_data_filename, 'r') as in_fp:
		with open(outfile, 'w') as out_fp:
			for line in in_fp:
				if first:
					print(line, file=out_fp)
					first = False
				else:
					fields = line.split('\t')
					# The key is (subject, group, ambiguity)
					key = (fields[4], fields[5], fields[6])
					record = keys_to_record[key]
					print('\t'.join([line.strip(), str(record[0]), str(record[1]), str(record[2]), str(record[3])]), file=out_fp)
